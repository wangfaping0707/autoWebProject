import openpyxl
"""
xlrd、xlwt和openpyxl模块的比较
1）xlrd：对xls、xlsx、xlsm文件进行读操作–读操作效率较高，推荐
2）xlwt：对xls文件进行写操作–写操作效率较高，但是不能执行xlsx文件
3）openpyxl：对xlsx、xlsm文件进行读、写操作–xlsx写操作推荐使用

1.2.2. 电子表格的基本概念
在使用Openpyxl前先要了解三个概念，即：Workbook(工作簿，一个包含多个Sheet的Excel文件)、
 Worksheet（工作表，一个Workbook有多个Worksheet，表名识别，如“Sheet1”,“Sheet2”等）、 Cell（单元格，存储具体的数据对象）三个对象。
在此将介绍使用 OpenpyXL 读取电子表格的方法。 OpenpyXL 使用下面的命令安装：
pip install Openpyxl
首先导入 openpyxl 模块，如果出错说明安装并未成功：
>>> import openpyxl
读取电子表格文件使用 openpyxl.load_workbook() 函数：
默认可读写，若有需要可以指定write_only和read_only为True 

1.2.3. 获取工作表
工作簿也就是我们能够在文件夹里看到的带有名字的Excel文件，当您双击这个文档时就打开了一个工作簿。 每一本工作簿可以拥有许多不同的工作表
在 Excel 中每次只能处理一个工作表，这个工作表的状态称为活动状态。 获得当前正在显示的工作表可以用 active 属性：
asheet = wb.active
每一个电子表格都有名称，通过 title 属性获得：
asheet.title
活动工作表的概念在使用 Excel 处理时比较方便。 但是编程处理中更倾向于使用更准确的方式来获取。 
要获得某个工作表， 可以根据工作表的名字获得。 要知道工作表的名字，
可以先通过 sheetnames 属性或 get_sheet_by_names() 函数把所有工作表的名称列出：
 wb.sheetnames
['中英文含摘要', '工作表2']
然后使用 get_sheet_by_name() 函数返回参数指定的工作表对象
sheet = wb.get_sheet_by_name('工作表2')
也可以通过索引值 ， 如下：
sheet2 = wb.worksheets[1]
工作表还有一些属性，如最大列数与最大行数，这两个数据表示工作表是实际存储数据的最大值，可以通过下面的属性获取：
sheet.max_row
sheet.max_column
根据工作表的最大列数与最大行数，可以用来对工作表的行、列进行遍历 。
"""
# 1、打开工作薄
wb = openpyxl.load_workbook(r'D:\webproject\utils\data.xlsx')
print(f'打印wb对象:{wb}')
# 2、获取工作薄中sheet表
# 当前正在操作的sheet表
s1 = wb.active
print(f'当前sheet表：{s1}')
# 获取工作薄中所有的sheet表名
print(f'工作薄中所有的sheet表:{wb.sheetnames}')
# 可以根据sheet表名称，指定要操作的sheet表
# sheet1 = wb.get_sheet_by_name('个人信息表')  这个方法被弃用了，使用时会报异常
sheet1 = wb['个人信息表']
print(f'打印sheet1：{sheet1}')
# 可以根据索引值，指定要操作的sheet表
sheet2 = wb.worksheets[1]
print(f'打印sheet2：{sheet2}')
print('===================================================')
# 获取sheet1表的最大行数
max_row =sheet1.max_row
print(f'sheet1表的最大行数:{max_row}')
# 获取sheet1表的最大列数
max_column = sheet1.max_column
print(f'sheet1表的最大列数:{max_column}')
print('===================================================')
"""
遍历行:我们可以使用iter_rows()方法来获取指定范围内的一系列行，并遍历每行中的每个单元格。
有 .iter_rows() 方式，肯定也会有 .iter_cols() 方式，只不过一个是按行读取，一个是按列读取
"""
# min_row 指定从那一行开始读取数据
# 按行读取数据
for row in sheet1.iter_rows(values_only=True,min_row=1,max_row=1):
	print(row)
print('==================================================')

# 按列读取数据
for col in sheet2.iter_cols(values_only=True,min_row=1):
	print(col)






































